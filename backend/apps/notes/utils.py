"""
Shared utilities for notes app - file parsing and column classification.
"""
import csv
import os
from dataclasses import dataclass
from typing import Optional, List, Any, Tuple
import openpyxl
import io


# Known module-level column suffixes (exact match after module code)
MODULE_COL_SUFFIXES = {'_Moy_Module', '_Note_Finale', '_Credit', '_Observation'}
# Known element-level column suffixes
ELEMENT_COL_SUFFIXES = {'_Continu', '_Final', '_Note', '_Credit', '_Obs'}

# Allowed file extensions for import
ALLOWED_EXTENSIONS = {'.csv', '.xlsx'}


@dataclass
class ColumnClassification:
    """Dataclass for column classification result."""
    col_type: str  # 'module' or 'element'
    module_code: str
    element_code: Optional[str] = None
    element_name: Optional[str] = None
    field: Optional[str] = None


def extract_code_and_name(col_name: str, suffix: str) -> Tuple[str, str]:
    """
    Extract code and name from column header.
    Format: {CODE}_{NAME}_{SUFFIX}
    Returns (code, name) where name has underscores replaced with spaces.
    """
    prefix = col_name[:-len(suffix)]
    # Find the first underscore to separate code from name
    first_underscore = prefix.find('_')
    if first_underscore == -1:
        return (prefix, '')
    code = prefix[:first_underscore]
    name = prefix[first_underscore + 1:].replace('_', ' ')
    return (code, name)


def classify_column(col_name: str, module_codes: set) -> Optional[ColumnClassification]:
    """
    Classify a column as module-level or element-level.
    Returns ColumnClassification or None.
    """
    for suffix in MODULE_COL_SUFFIXES:
        if col_name.endswith(suffix):
            module_code, module_name = extract_code_and_name(col_name, suffix)
            if module_code in module_codes:
                field = suffix[1:]  # Remove leading underscore
                return ColumnClassification(
                    col_type='module',
                    module_code=module_code,
                    element_code=None,
                    element_name=module_name,
                    field=field
                )

    for suffix in ELEMENT_COL_SUFFIXES:
        if col_name.endswith(suffix):
            element_code, element_name = extract_code_and_name(col_name, suffix)
            # Check if this element_code starts with any known module code
            for mc in module_codes:
                if element_code.startswith(mc) and len(element_code) > len(mc):
                    return ColumnClassification(
                        col_type='element',
                        module_code=mc,
                        element_code=element_code,
                        element_name=element_name,
                        field=suffix[1:]
                    )

    return None


def extract_module_codes(headers: List[str]) -> set:
    """Extract module codes from column headers."""
    codes = set()
    for col in headers:
        for suffix in MODULE_COL_SUFFIXES:
            if col.endswith(suffix):
                # Extract just the code (first segment before underscore)
                prefix = col[:-len(suffix)]
                first_underscore = prefix.find('_')
                if first_underscore != -1:
                    codes.add(prefix[:first_underscore])
                else:
                    codes.add(prefix)
    return codes


def parse_and_validate_file(fichier) -> Tuple[List[str], List[Any]]:
    """
    Parse and validate the uploaded file.
    Returns (headers, rows) or raises ValueError for invalid format.
    """
    # Validate file extension
    file_ext = os.path.splitext(fichier.name)[1].lower()
    if file_ext not in ALLOWED_EXTENSIONS:
        raise ValueError(f"Extension de fichier non autorisée. Utilisez {', '.join(ALLOWED_EXTENSIONS)}")
    
    if file_ext == '.xlsx':
        wb = openpyxl.load_workbook(fichier)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = [
            dict(zip(headers, row))
            for row in ws.iter_rows(min_row=2, values_only=True)
        ]
    else:  # .csv
        decoded = fichier.read().decode('utf-8')
        reader = csv.DictReader(io.StringIO(decoded))
        headers = reader.fieldnames or []
        rows = list(reader)
    
    return headers, rows