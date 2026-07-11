"""
Admin views for import and reporting.
"""
from rest_framework import generics, permissions, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from django.db import transaction
import openpyxl
import csv
import io
from apps.notes.models import ResultatSemestre, Module, ElementModule
from apps.users.models import User
from .permissions import IsAdmin


# Known module-level column suffixes (exact match after module code)
MODULE_COL_SUFFIXES = {'_Moy_Module', '_Note_Finale', '_Credit', '_Observation'}
# Known element-level column suffixes
ELEMENT_COL_SUFFIXES = {'_Continu', '_Final', '_Note', '_Credit', '_Obs'}


def classify_column(col_name, module_codes):
    """
    Classify a column as module-level or element-level.
    Returns ('module', module_code, field) or ('element', element_code, field) or None.
    """
    for suffix in MODULE_COL_SUFFIXES:
        if col_name.endswith(suffix):
            module_code = col_name[:-len(suffix)]
            if module_code in module_codes:
                field = suffix[1:]  # Remove leading underscore
                return ('module', module_code, field)

    for suffix in ELEMENT_COL_SUFFIXES:
        if col_name.endswith(suffix):
            element_code = col_name[:-len(suffix)]
            # Check if this element_code starts with any known module code
            for mc in module_codes:
                if element_code.startswith(mc) and len(element_code) > len(mc):
                    return ('element', mc, element_code, suffix[1:])

    return None


def extract_module_codes(headers):
    """Extract module codes from column headers."""
    codes = set()
    for col in headers:
        for suffix in MODULE_COL_SUFFIXES:
            if col.endswith(suffix):
                codes.add(col[:-len(suffix)])
    return codes


@api_view(['POST'])
@permission_classes([IsAdmin])
def import_pv(request):
    """
    POST /api/admin/import-pv/
    Admin: bulk import ResultatSemestre/Module/ElementModule from CSV/XLSX file.
    Expected format: Numéro, Nom et Prénom, Moy. Semestre, Observation,
    then for each module: MIAG31_Moy_Module, MIAG31_Note_Finale, MIAG31_Credit, MIAG31_Observation,
    then for each element: MIAG311_Continu, MIAG311_Final, MIAG311_Note, MIAG311_Credit, MIAG311_Obs
    """
    fichier = request.FILES.get('fichier')
    if not fichier:
        return Response(
            {"detail": "Fichier requis."},
            status=status.HTTP_400_BAD_REQUEST
        )

    imported_count = 0
    errors = []
    semestre = request.POST.get('semestre', 'S3')
    annee_academique = request.POST.get('annee_academique', '2024-2025')

    try:
        if fichier.name.endswith('.xlsx'):
            wb = openpyxl.load_workbook(fichier)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        elif fichier.name.endswith('.csv'):
            decoded = fichier.read().decode('utf-8')
            reader = csv.DictReader(io.StringIO(decoded))
            rows = list(reader)
            headers = list(reader.fieldnames) if reader.fieldnames else []
        else:
            return Response(
                {"detail": "Format non supporté. Utilisez CSV ou XLSX."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Extract module codes from headers
        module_codes = extract_module_codes(headers)

        with transaction.atomic():
            for row in rows:
                try:
                    # Handle both dict and tuple formats
                    if isinstance(row, dict):
                        matricule = row.get('Numéro', '').strip()
                        nom_prenom = row.get('Nom et Prénom', '').strip()
                        moy_semestre = row.get('Moy. Semestre', '0').strip()
                        observation = row.get('Observation', 'Rattrapage').strip()
                    else:
                        matricule = str(row[0]).strip() if row[0] else ''
                        nom_prenom = str(row[1]).strip() if row[1] else ''
                        moy_semestre = str(row[2]).strip() if row[2] else '0'
                        observation = str(row[3]).strip() if row[3] else 'Rattrapage'

                    if not matricule:
                        errors.append(f"Ligne invalide: {row}")
                        continue

                    try:
                        etudiant, created = User.objects.get_or_create(
                            matricule=matricule,
                            defaults={
                                'email': f"{matricule.lower()}@iscae.edu.mr",
                                'role': 'ETUDIANT',
                            }
                        )
                        if created and nom_prenom:
                            names = nom_prenom.split()
                            if len(names) >= 2:
                                etudiant.last_name = ' '.join(names[:-1])
                                etudiant.first_name = names[-1]
                            else:
                                etudiant.last_name = nom_prenom
                            etudiant.save()
                    except Exception as e:
                        errors.append(f"Erreur étudiant {matricule}: {str(e)}")
                        continue

                    # Create/update ResultatSemestre
                    resultat, _ = ResultatSemestre.objects.update_or_create(
                        etudiant=etudiant,
                        semestre=semestre,
                        annee_academique=annee_academique,
                        defaults={
                            'moy_semestre': float(moy_semestre) if moy_semestre else 0,
                            'observation': observation,
                        }
                    )

                    # First pass: collect module-level data
                    module_data = {}
                    for col_name, value in row.items():
                        if col_name in ['Numéro', 'Nom et Prénom', 'Moy. Semestre', 'Observation']:
                            continue
                        classification = classify_column(col_name, module_codes)
                        if classification and classification[0] == 'module':
                            _, mc, field = classification
                            if mc not in module_data:
                                module_data[mc] = {}
                            module_data[mc][field] = value

                    # Second pass: collect element-level data per module
                    element_data = {}
                    for col_name, value in row.items():
                        if col_name in ['Numéro', 'Nom et Prénom', 'Moy. Semestre', 'Observation']:
                            continue
                        classification = classify_column(col_name, module_codes)
                        if classification and classification[0] == 'element':
                            _, mc, ec, field = classification
                            key = (mc, ec)
                            if key not in element_data:
                                element_data[key] = {}
                            element_data[key][field] = value

                    # Create/update modules
                    for mc, mdata in module_data.items():
                        module, _ = Module.objects.update_or_create(
                            resultat_semestre=resultat,
                            code_module=mc,
                            defaults={
                                'moy_module': float(mdata.get('moy_module', 0) or 0),
                                'note_finale': float(mdata.get('note_finale', 0) or 0),
                                'credit': float(mdata.get('credit', 0) or 0),
                                'observation': mdata.get('observation', 'Rattrapage'),
                            }
                        )

                        # Create/update elements for this module
                        for (mc_key, ec), edata in element_data.items():
                            if mc_key != mc:
                                continue

                            ElementModule.objects.update_or_create(
                                module=module,
                                code_element=ec,
                                defaults={
                                    'note_continu': float(edata.get('Continu', 0) or 0),
                                    'note_final': float(edata.get('Final', 0) or 0),
                                    'note_moyenne': float(edata.get('Note', 0) or 0),
                                    'credit': float(edata.get('Credit', 0) or 0),
                                    'observation': edata.get('Obs', 'Rattrapage'),
                                }
                            )
                            imported_count += 1

                except Exception as e:
                    errors.append(f"Erreur sur ligne {row}: {str(e)}")

        return Response({
            "detail": f"Import terminé. {imported_count} élément(s) importé(s).",
            "imported": imported_count,
            "errors": errors[:50],
        })

    except Exception as e:
        return Response(
            {"detail": f"Erreur lors de l'import: {str(e)}"},
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['POST'])
@permission_classes([IsAdmin])
def force_unblock_note(request, reclamation_id):
    """
    POST /api/admin/reclamations/{id}/force-unblock/
    RG-03: Admin can force-unblock a previously accepted element,
    allowing a new reclamation to be submitted.
    Archives the accepted reclamation to remove the block.
    """
    from .models import Reclamation, StatutReclamation, HistoriqueStatut

    try:
        reclamation = Reclamation.objects.get(id=reclamation_id)
    except Reclamation.DoesNotExist:
        return Response(
            {"detail": "Réclamation introuvable."},
            status=status.HTTP_404_NOT_FOUND
        )

    if reclamation.statut != StatutReclamation.ACCEPTEE:
        return Response(
            {"detail": "Seules les réclamations acceptées peuvent être débloquées."},
            status=status.HTTP_400_BAD_REQUEST
        )

    with transaction.atomic():
        old_statut = reclamation.statut
        reclamation.statut = StatutReclamation.ARCHIVEE
        reclamation.save(update_fields=['statut'])

        HistoriqueStatut.objects.create(
            reclamation=reclamation,
            statut_precedent=old_statut,
            nouveau_statut=StatutReclamation.ARCHIVEE,
            commentaire="Déblocage forcé par l'administrateur (RG-03)",
            modifie_par=request.user,
        )

    return Response({
        "detail": f"Réclamation #{reclamation_id} débloquée. L'étudiant peut soumettre une nouvelle réclamation.",
        "reclamation_id": reclamation.id,
        "nouveau_statut": StatutReclamation.ARCHIVEE,
    })


@api_view(['GET'])
@permission_classes([IsAdmin])
def export_rapport(request):
    """
    GET /api/rapports/?format=csv
    Admin: export reclamation statistics.
    """
    from .models import Reclamation
    from django.http import HttpResponse

    format_type = request.query_params.get('format', 'csv')

    reclamations = Reclamation.objects.select_related(
        'etudiant', 'coordonnateur'
    ).prefetch_related('lignes', 'lignes__element_module', 'lignes__element_module__module').all()

    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="rapport_reclamations.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'ID', 'Matricule', 'Étudiant', 'Module', 'Élément', 'Type', 'Motif', 'Statut',
            'Date création', 'Date limite', 'Date traitement',
            'Note originale', 'Nouvelle note', 'Coordinateur', 'Commentaire'
        ])

        for r in reclamations:
            for l in r.lignes.all():
                writer.writerow([
                    r.id,
                    r.etudiant.matricule,
                    r.etudiant.get_full_name(),
                    l.element_module.module.code_module if l.element_module else '',
                    l.element_module.code_element if l.element_module else '',
                    l.get_type_note_display(),
                    l.motif,
                    r.statut,
                    r.date_creation.strftime('%Y-%m-%d %H:%M'),
                    r.date_limite_traitement.strftime('%Y-%m-%d %H:%M'),
                    r.date_traitement.strftime('%Y-%m-%d %H:%M') if r.date_traitement else '',
                    l.note_originale,
                    l.nouvelle_note,
                    r.coordonnateur.get_full_name() if r.coordonnateur else '',
                    r.commentaire_decision,
                ])

        return response

    return Response({"detail": "Format non supporté. Utilisez ?format=csv"})